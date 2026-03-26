import os
import io
import json
import openpyxl
import datetime
import shutil
import tempfile
import subprocess
from collections import defaultdict
import math
import re
import glob
import pandas as pd
from flask import Flask, render_template, jsonify, request, send_file
from jinja2 import ChoiceLoader, FileSystemLoader
from dotenv import load_dotenv
from supabase import create_client, Client
from dateutil.parser import parse as date_parse

load_dotenv()
SUPABASE_URL = os.getenv('SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_KEY')

app = Flask(__name__)

# 파일 경로 스키마 정의 (Vercel Serverless 호환되도록 동적 경로 사용)
# Vercel에서는 /var/task 등 동적 디렉토리에서 실행되므로 __file__ 기반 절대경로 추출이 필수
if os.environ.get('VERCEL'):
    # Vercel 환경: api 폴더에서 실행될 경우 프로젝트 루트 지정 등 필요 시 조정
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
else:
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

print(f"--- SERVER STARTING ---")
print(f"BASE_DIR: {BASE_DIR}")
print(f"CWD: {os.getcwd()}")
print(f"-----------------------")

# 템플릿 검색 경로 설정 (루트 및 web/templates 동시 탐색)
app.jinja_loader = ChoiceLoader([
    FileSystemLoader(BASE_DIR),
    FileSystemLoader(os.path.join(BASE_DIR, 'web', 'templates'))
])

# 서버리스에서는 /tmp 디렉토리만 쓰기가 가능하지만, 여기선 단순 읽기 목적이므로 프로젝트 내 파일 참조
# find_data_file 정의 후 아래에서 설정됨 (L80 부근)
JSON_FILE = ""
THRESHOLD_FILE = ""

def get_supabase_client() -> Client:
    if not SUPABASE_URL or not SUPABASE_KEY:
        raise Exception("Supabase 환경 변수가 설정되지 않았습니다.")
    return create_client(SUPABASE_URL, SUPABASE_KEY)

def find_data_file(relative_path):
    """데이터 파일의 실제 위치를 찾는 헬퍼 함수 (로컬 및 Vercel 대응)"""
    # 탐색할 후보 경로들
    candidates = [
        # 1. 원래 루트 기준 경로 (로컬)
        os.path.normpath(os.path.join(BASE_DIR, 'data', relative_path)),
        # 2. Vercel용 api 하위 경로
        os.path.normpath(os.path.join(BASE_DIR, 'api', 'data', relative_path)),
        # 3. 현재 파일 기준 상대 경로 (web/app.py 기준)
        os.path.normpath(os.path.join(os.path.dirname(__file__), '..', 'data', relative_path)),
        # 4. 현재 작업 디렉토리 기준
        os.path.normpath(os.path.join(os.getcwd(), 'data', relative_path)),
        os.path.normpath(os.path.join(os.getcwd(), 'api', 'data', relative_path)),
    ]
    
    for p in candidates:
        if os.path.exists(p):
            return p
            
    # 기본값 반환 하되 로그를 남김
    print(f"Warning: Data file NOT found: {relative_path}. Tried: {candidates}")
    return candidates[0]

# 서버리스에서는 /tmp 디렉토리만 쓰기가 가능하지만, 여기선 단순 읽기 목적이므로 프로젝트 내 파일 참조
JSON_FILE = find_data_file('material_master.json')
THRESHOLD_FILE = find_data_file('inventory_thresholds.json')

@app.route('/')
def index():
    try:
        return render_template('index.html')
    except Exception as e:
        import traceback
        trace_str = traceback.format_exc()
        html_files = []
        if os.path.exists(BASE_DIR):
            html_files = [f for f in os.listdir(BASE_DIR) if f.endswith('.html')]
        return f"Error: {str(e)}<br>BASE_DIR: {BASE_DIR}<br>Files in BASE_DIR: {html_files}<br><pre>{trace_str}</pre>", 500

@app.route('/inventory')
def inventory():
    return render_template('inventory.html')

@app.route('/production')
def production():
    return render_template('production.html')

@app.route('/producible')
def producible():
    return render_template('producible.html')

@app.route('/material-info')
def material_info():
    return render_template('material_info.html')

@app.route('/upload-usage')
def upload_usage():
    return render_template('upload_usage.html')

@app.route('/upload-receiving')
def upload_receiving():
    return render_template('upload_receiving.html')

@app.route('/finished-product')
def finished_product():
    return render_template('finished_product.html')

@app.route('/raw-material')
def raw_material():
    return render_template('raw_material.html')

@app.route('/validation')
def validation():
    return render_template('validation.html')

@app.route('/calibration')
def calibration():
    return render_template('calibration.html')

@app.route('/purchase-dashboard')
def purchase_dashboard():
    return render_template('purchase_dashboard.html')

@app.route('/facilities')
def facilities():
    return render_template('facilities.html')

@app.route('/mo-management')
def mo_management():
    return render_template('mo_management.html')

def get_red_cells_data(file_path):
    if not os.path.exists(file_path): return []
    wb = openpyxl.load_workbook(file_path, data_only=True)
    red_cells = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and cell.font and cell.font.color:
                    color = str(cell.font.color.rgb)
                    # 빨간색 (Standard Red or ARGB FF0000)
                    if color == 'FFFF0000' or color == 'FF0000' or (isinstance(color, str) and color.endswith('0000') and color.startswith('FF')):
                        label = f"Cell {cell.coordinate}"
                        left_cell = sheet.cell(row=cell.row, column=max(1, cell.column - 1))
                        if left_cell.value and isinstance(left_cell.value, str):
                            label = left_cell.value
                        else:
                            above_cell = sheet.cell(row=max(1, cell.row - 1), column=cell.column)
                            if above_cell.value and isinstance(above_cell.value, str):
                                label = above_cell.value

                        val = cell.value
                        if isinstance(val, datetime.datetime):
                            val = val.strftime("%Y-%m-%d")

                        red_cells.append({
                            'sheet': sheet_name,
                            'cell': cell.coordinate,
                            'label': label.strip() if isinstance(label, str) else str(label),
                            'value': val
                        })
    return red_cells

@app.route('/api/mo/red-cells')
def api_mo_red_cells():
    base_excel = find_data_file(os.path.join('mo', 'MO_RESULT.xlsx'))
    red_cells = get_red_cells_data(base_excel)
    return jsonify({'status': 'success', 'data': red_cells})

@app.route('/api/mo/generate', methods=['POST'])
def api_mo_generate():
    base_excel = find_data_file(os.path.join('mo', 'MO_RESULT.xlsx'))
    
    # recalc.py 경로 수정 (api 폴더 내에 복사된 버전 사용)
    # Vercel 환경에서는 BASE_DIR/api/recalc.py 또는 루트/api/recalc.py 등을 탐색
    candidates_recalc = [
        os.path.join(BASE_DIR, 'data', 'recalc.py'),
        os.path.join(BASE_DIR, 'api', 'recalc.py'),
        os.path.join(os.path.dirname(__file__), '..', 'data', 'recalc.py'),
        os.path.join(os.getcwd(), 'data', 'recalc.py')
    ]
    recalc_script = next((p for p in candidates_recalc if os.path.exists(p)), candidates_recalc[0])
    
    if not os.path.exists(base_excel):
        # 상세한 디버깅 정보 포함
        available_files = []
        try:
            for root, dirs, files in os.walk(BASE_DIR):
                for f in files:
                    if f.endswith('.xlsx'):
                        available_files.append(os.path.relpath(os.path.join(root, f), BASE_DIR))
        except: pass

        return jsonify({
            'status': 'error', 
            'message': f'Base MO Excel file not found. Checked multiple locations including {base_excel}',
            'cwd': os.getcwd(),
            'base_dir': BASE_DIR,
            'available_xlsx': available_files[:10] # 너무 많을 수 있으니 일부만
        }), 404

    # 임시 파일 경로 생성
    fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    
    try:
        shutil.copyfile(base_excel, temp_path)
        wb = openpyxl.load_workbook(temp_path)
        
        # 폼 데이터 처리
        for key, value in request.form.items():
            if "_" in key:
                try:
                    sheet_name, cell_coord = key.split("_", 1)
                    if sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        
                        # 데이터 타입 변환 시도
                        processed_value = value
                        if value:
                            try:
                                if "." in value:
                                    processed_value = float(value)
                                else:
                                    processed_value = int(value)
                            except ValueError:
                                try:
                                    processed_value = date_parse(value)
                                except:
                                    processed_value = value
                        
                        sheet[cell_coord] = processed_value
                except Exception as e:
                    print(f"Error processing {key}: {e}")

        wb.save(temp_path)

        # 수식 재계산 (LibreOffice 활용)
        if os.path.exists(recalc_script):
            try:
                subprocess.run(["python", recalc_script, temp_path], capture_output=True, text=True, check=False, timeout=60)
            except Exception as e:
                print(f"Recalculation error: {e}")
        
        output_filename = f"MO_RESULT_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(temp_path, as_attachment=True, download_name=output_filename)

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        # 파일이 바로 삭제되면 send_file에서 문제가 생길 수 있으므로 주의 필요하지만 
        # Flask 2.2+ 혹은 as_attachment 시 파일을 메모리에 로드 후 전송하므로 안전할 가능성이 큼
        # 다만 좀 더 안전하게 삭제하려면 별도 로직이 필요할 수 있음
        pass

@app.route('/api/debug/files')
def debug_files():
    files_tree = []
    try:
        # 주요 디렉토리 탐색
        for d in [BASE_DIR, os.getcwd(), os.path.join(BASE_DIR, 'data'), os.path.join(BASE_DIR, 'api')]:
            if os.path.exists(d):
                for root, dirs, files in os.walk(d):
                    # 너무 많은 파일 방지를 위해 depth 제한 (간접적으로)
                    level = root.replace(d, '').count(os.sep)
                    if level > 3: continue
                    
                    for file in files:
                        files_tree.append({
                            'path': os.path.relpath(os.path.join(root, file), BASE_DIR),
                            'abs': os.path.join(root, file),
                            'size': os.path.getsize(os.path.join(root, file))
                        })
    except Exception as e:
        return jsonify({'error': str(e)})

    return jsonify({
        'base_dir': BASE_DIR,
        'cwd': os.getcwd(),
        'env_vercel': os.environ.get('VERCEL'),
        'files': files_tree
    })

@app.route('/api/purchase/info')
def get_purchase_info():
    try:
        supabase = get_supabase_client()
        response = supabase.table('purchase_info').select('*').execute()
        return jsonify({'status': 'success', 'data': response.data})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/validation/plan')
def get_validation_plan():
    try:
        supabase = get_supabase_client()
        response = supabase.table('validation_plan').select('*').execute()
        # 번호순 정렬 (숫자로 변환하여 정렬)
        data = response.data
        data.sort(key=lambda x: int(x['no']) if str(x['no']).isdigit() else 9999)
        return jsonify({"status": "success", "data": data})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/calibration/plan')
def get_calibration_plan():
    try:
        supabase = get_supabase_client()
        response = supabase.table('instrument_calibration').select('*').execute()
        # 번호순 정렬
        data = response.data
        data.sort(key=lambda x: int(x['no']) if str(x['no']).isdigit() else 9999)
        return jsonify({"status": "success", "data": data})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/facilities')
def get_facilities():
    try:
        supabase = get_supabase_client()
        response = supabase.table('facilities').select('*').order('id').execute()
        return jsonify({"status": "success", "data": response.data})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/producible')
def get_producible():
    """현 재고량 기준 원료별 최대 생산가능 kit수 계산"""
    try:
        import bisect

        # ── 1. BOM.xlsx 파싱 ──────────────────────────────────────
        BOM_FILE = find_data_file('BOM.xlsx')
        wb = openpyxl.load_workbook(BOM_FILE, read_only=True, data_only=True)
        ws = wb.active

        header = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        kit_cols = len(header) - 2  # 256

        bom = {}
        bom_name = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            code = row[0]
            if not code:
                continue
            name = row[1] or ''
            usages = [row[2 + i] or 0 for i in range(kit_cols)]
            bom[code] = usages
            bom_name[code] = name
        wb.close()

        # ── 2. 단위 매핑 ────────────────────────────────────────
        unit_map = {}
        try:
            with open(JSON_FILE, 'r', encoding='utf-8') as f:
                master = json.load(f)
            unit_map = {code: info.get('단위', '') for code, info in master.items()}
        except Exception:
            pass

        # ── 3. DB 현재고 조회 (Supabase) ─────────────────────────
        sb = get_supabase_client()
        # 모든 raw_materials 가져오기
        res = sb.table('raw_materials').select('item_code, product_name, transaction_type, quantity').execute()
        
        stock_map = defaultdict(lambda: {'product_name': '', 'total_stock': 0.0})
        for r in res.data:
            code = r['item_code']
            qty = r['quantity'] or 0
            if r['transaction_type'] == '입고':
                stock_map[code]['total_stock'] += qty
            else:
                stock_map[code]['total_stock'] -= qty
            if r.get('product_name'):
                stock_map[code]['product_name'] = r['product_name']

        # ── 4. 계산 ───────────────────
        results = []
        for code, usages in bom.items():
            stock_info = stock_map.get(code, {})
            current_stock = stock_info.get('total_stock', 0)
            product_name = stock_info.get('product_name') or bom_name.get(code, '')

            max_kit = 0
            for i, usage in enumerate(usages):
                if usage is None or usage == 0:
                    max_kit = kit_cols
                elif current_stock >= usage:
                    max_kit = i + 1
                else:
                    break

            results.append({
                'item_code': code,
                'product_name': product_name,
                'current_stock': round(current_stock, 4),
                'unit': unit_map.get(code, ''),
                'max_kit': max_kit,
            })

        results.sort(key=lambda x: x['item_code'])
        return jsonify({'status': 'success', 'data': results})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/materials')
def get_materials():
    """material_master.json에서 원료 코드와 이름을 가져옵니다."""
    try:
        with open(JSON_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
            
        materials = []
        for code, details in data.items():
            materials.append({
                'code': code,
                'name': details.get('제품명', 'Unknown')
            })
            
        materials = sorted(materials, key=lambda x: x['code'])
        return jsonify({'status': 'success', 'data': materials})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/lots/<item_code>')
def get_lots(item_code):
    try:
        sb = get_supabase_client()
        res = sb.table('raw_materials').select('lot_no').eq('item_code', item_code).execute()
        lots = list(set([r['lot_no'] for r in res.data if r.get('lot_no')]))
        lots.sort()
        return jsonify({'status': 'success', 'data': lots})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/inventory')
def get_inventory():
    item_code = request.args.get('item_code')
    lot_no = request.args.get('lot_no')
    
    if not item_code or not lot_no:
        return jsonify({'status': 'error', 'message': 'Missing item_code or lot_no'}), 400
        
    try:
        sb = get_supabase_client()
        # 이력 조회
        res_hist = sb.table('raw_materials').select('id, product_name, transaction_type, transaction_date, purpose, quantity').eq('item_code', item_code).eq('lot_no', lot_no).order('transaction_date').order('id').execute()
        history = res_hist.data
        
        # 원료 상세 조회
        res_info = sb.table('material_info').select('receive_date, qc_date, expire_date').eq('item_code', item_code).eq('lot_no', lot_no).limit(1).execute()
        material_details = res_info.data[0] if res_info.data else None
        
        total_in = sum(item['quantity'] for item in history if item['transaction_type'] == '입고')
        total_out = sum(item['quantity'] for item in history if item['transaction_type'] == '출고')
        current_stock = total_in - total_out
        
        summary = {
            'total_in': total_in,
            'total_out': total_out,
            'current_stock': current_stock
        }
        
        return jsonify({
            'status': 'success', 
            'summary': summary,
            'history': history,
            'material_details': material_details
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/stock/summary')
def get_stock_summary():
    try:
        unit_map = {}
        try:
            with open(JSON_FILE, 'r', encoding='utf-8') as f:
                master = json.load(f)
            unit_map = {code: info.get('단위', '') for code, info in master.items()}
        except Exception:
            pass

        sb = get_supabase_client()
        res = sb.table('raw_materials').select('item_code, product_name, transaction_type, quantity').execute()
        
        stats = defaultdict(lambda: {'product_name': '', 'total_in': 0.0, 'total_out': 0.0})
        for r in res.data:
            code = r['item_code']
            qty = r['quantity'] or 0
            if r['transaction_type'] == '입고':
                stats[code]['total_in'] += qty
            else:
                stats[code]['total_out'] += qty
            if r.get('product_name'):
                stats[code]['product_name'] = r['product_name']

        rows = []
        for code, data in stats.items():
            rows.append({
                'item_code': code,
                'product_name': data['product_name'],
                'total_stock': data['total_in'] - data['total_out'],
                'unit': unit_map.get(code, '')
            })
            
        rows.sort(key=lambda x: x['item_code'])
        return jsonify({'status': 'success', 'data': rows})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/material-info')
def get_material_info():
    try:
        with open(JSON_FILE, 'r', encoding='utf-8') as f:
            master = json.load(f)

        thresholds = {}
        try:
            with open(THRESHOLD_FILE, 'r', encoding='utf-8') as f:
                thresholds = json.load(f)
        except Exception:
            pass

        sb = get_supabase_client()
        res = sb.table('raw_materials').select('item_code, transaction_type, quantity').execute()
        
        stock_map = defaultdict(float)
        for r in res.data:
            qty = r['quantity'] or 0
            if r['transaction_type'] == '입고':
                stock_map[r['item_code']] += qty
            else:
                stock_map[r['item_code']] -= qty

        result = []
        for code, info in master.items():
            safe = thresholds.get(code, {}).get('safe_stock_level', None)
            current = stock_map.get(code, 0.0)

            result.append({
                'item_code': code,
                'product_name': info.get('제품명', ''),
                'cat_no': info.get('Cat_No', ''),
                'package_unit': info.get('포장단위', ''),
                'unit': info.get('단위', ''),
                'manufacturer': info.get('제조사', ''),
                'storage_temp': info.get('보관온도', ''),
                'storage_location': info.get('보관장소', ''),
                'unit_price': info.get('단가', None),
                'safe_stock': safe,
                'current_stock': round(current, 4)
            })

        result.sort(key=lambda x: x['item_code'])
        return jsonify({'status': 'success', 'data': result})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/production/calculate')
def calculate_production():
    kit_qty = request.args.get('kit_qty', type=int)
    if not kit_qty or kit_qty < 1 or kit_qty > 256:
        return jsonify({'status': 'error', 'message': 'Invalid kit quantity (1-256)'}), 400
        
    try:
        sb = get_supabase_client()
        # 1. BOM 소요량 조회
        res_bom = sb.table('kit_bom').select('material_code, material_name, usage_qty').eq('kit_qty', kit_qty).execute()
        bom_items = res_bom.data

        # 2. 모든 material_info (rec_date, expire_date용)
        res_info = sb.table('material_info').select('item_code, lot_no, receive_date, expire_date').execute()
        info_map = {}
        for r in res_info.data:
            info_map[(r['item_code'], r['lot_no'])] = {
                'receive_date': r['receive_date'],
                'expire_date': r['expire_date']
            }

        # 3. 모든 raw_materials 재고 조회 (각 lot 별 잔량 계산)
        codes = [item['material_code'] for item in bom_items]
        res_raw = sb.table('raw_materials').select('item_code, lot_no, transaction_type, quantity').in_('item_code', codes).execute()
        
        lot_stocks = defaultdict(float)
        lot_in_qty = defaultdict(float)
        for r in res_raw.data:
            key = (r['item_code'], r['lot_no'])
            if r['transaction_type'] == '입고':
                lot_stocks[key] += (r['quantity'] or 0)
                lot_in_qty[key] += (r['quantity'] or 0)
            else:
                lot_stocks[key] -= (r['quantity'] or 0)

        results = []
        for item in bom_items:
            mat_code = item['material_code']
            mat_name = item['material_name']
            required_qty = item['usage_qty']

            # 가용 재고 목록 생성
            available_lots = []
            for (itm_cd, lot_no), stock in lot_stocks.items():
                # 부동 소수점 오차를 고려하여 0.0001보다 큰 경우만 가용 재고로 인정
                if itm_cd == mat_code and stock > 0.0001:
                    info = info_map.get((itm_cd, lot_no), {})
                    rec_date = info.get('receive_date') or '9999-12-31'
                    exp_date = info.get('expire_date') or '9999-12-31'
                    
                    # 입고량 넷 대비 잔량이 적으면 이미 사용중인(뜯은) 원료로 간주
                    total_in = lot_in_qty.get((itm_cd, lot_no), 0)
                    is_in_use = 0 if stock < total_in else 1
                    
                    available_lots.append({
                        'lot_no': lot_no,
                        'current_stock': stock,
                        'rec_date': rec_date,
                        'exp_date': exp_date,
                        'is_in_use': is_in_use
                    })
            
            # 최우선: 사용중인(0) 로트 -> 유효기간 빠른 순 -> 입고일 빠른 순 정렬
            available_lots.sort(key=lambda x: (x['is_in_use'], x['exp_date'], x['rec_date'], x['lot_no']))

            allocated_lots = []
            remaining_to_allocate = required_qty
            
            for lot in available_lots:
                if remaining_to_allocate <= 0:
                    break
                    
                alloc_qty = min(remaining_to_allocate, lot['current_stock'])
                allocated_lots.append({
                    'lot_no': lot['lot_no'],
                    'receive_date': lot['rec_date'] if lot['rec_date'] != '9999-12-31' else '정보 없음',
                    'expire_date': lot['exp_date'] if lot['exp_date'] != '9999-12-31' else '정보 없음',
                    'available_stock': lot['current_stock'],
                    'allocated_qty': round(alloc_qty, 4)
                })
                
                remaining_to_allocate -= alloc_qty
                lot_stocks[(mat_code, lot['lot_no'])] = max(0, lot_stocks[(mat_code, lot['lot_no'])] - alloc_qty)
                
            shortage_qty = round(max(0, remaining_to_allocate), 4)
            
            results.append({
                'material_code': mat_code,
                'material_name': mat_name,
                'required_qty': round(required_qty, 4),
                'allocated_lots': allocated_lots,
                'shortage_qty': shortage_qty,
                'status': 'success' if shortage_qty == 0 else 'shortage'
            })
            
        return jsonify({'status': 'success', 'data': results})
    except Exception as e:
        import traceback
        err_msg = traceback.format_exc()
        print(f"ERROR in get_finished_product_statistics: {err_msg}")
        return jsonify({'status': 'error', 'message': str(e), 'trace': err_msg}), 500

@app.route('/api/receiving/upload', methods=['POST'])
def upload_receiving_api():
    if 'file' not in request.files:
        return jsonify({'status': 'error', 'message': '파일이 첨부되지 않았습니다.'}), 400

    file = request.files['file']
    if not file.filename or not file.filename.endswith('.xlsx'):
        return jsonify({'status': 'error', 'message': '.xlsx 파일만 업로드 가능합니다.'}), 400

    try:
        # 단일 시트 데이터 로드
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), read_only=True, data_only=True)
        ws = wb.active

        # 지정된 셀 좌표에서 값 추출
        def gval(cell):
            val = ws[cell].value
            if val is None:
                return ''
            if hasattr(val, 'strftime'):  # datetime 처리
                return val.strftime('%Y-%m-%d')
            return str(val).strip()

        # 사용자 요청 엑셀 좌표 업데이트
        product_name = gval('D5')
        manufacturer = gval('D6')
        quantity_raw = gval('D7')
        po_no = gval('D9')
        
        cat_no = gval('H5')
        vendor = gval('H6')
        lot_no = gval('H7')
        
        receive_date = gval('O5')
        qc_date = gval('O6')
        expire_date = gval('O7')
        item_code = gval('O8')
        storage_location = gval('H8') # 스크린샷 참고 (보관장소 위치 추정)
        
        wb.close()

        # 데이터 유효성 검사
        if not item_code or not lot_no:
            return jsonify({'status': 'error', 'message': '엑셀 양식 오류: 제품 코드(O8) 또는 Lot No(H7)가 누락되었습니다.'}), 400
            
        import re
        # 수량에서 숫자만 추출 (ex: 500g -> 500)
        quantity = 0.0
        match = re.search(r'[\d\.]+', quantity_raw)
        if match:
            quantity = float(match.group())
        if quantity <= 0:
            return jsonify({'status': 'error', 'message': f'유효하지 않은 입고수량입니다. (추출값: {quantity_raw})'}), 400

        sb = get_supabase_client()
        
        # 1단계: material_info 테이블 업데이트/삽입 (수동 upsert 로직 + ID 자동 할당)
        try:
            # 기존 레코드 존재 여부 확인 (item_code와 lot_no 기준)
            existing = sb.table('material_info').select('id').eq('item_code', item_code).eq('lot_no', lot_no).execute()
            
            info_data = {
                'item_code': item_code,
                'product_name': product_name,
                'cat_no': cat_no,
                'lot_no': lot_no,
                'purchase_qty': quantity,
                'manufacturer': manufacturer,
                'vendor': vendor,
                'receive_date': receive_date if receive_date else None,
                'qc_date': qc_date if qc_date else None,
                'expire_date': expire_date if expire_date else None,
                'po_no': po_no
            }

            if existing.data:
                # 이미 존재하면 업데이트
                row_id = existing.data[0]['id']
                sb.table('material_info').update(info_data).eq('id', row_id).execute()
                print(f"material_info updated (ID: {row_id})")
            else:
                # 존재하지 않으면 삽입 (DB가 ID 자동 할당)
                sb.table('material_info').insert(info_data).execute()
                print(f"material_info inserted (auto-id)")
        except Exception as e_info:
            print(f"material_info processing error: {str(e_info)}")
            # 정보 테이블 저장 실패 시에도 로그를 남기고 계속 진행 (또는 에러 반환 선택 가능)

        # 2단계: raw_materials 입고 내역 추가 (ID 자동 할당)
        try:
            sb.table('raw_materials').insert({
                'item_code': item_code,
                'product_name': product_name if product_name else '알 수 없음',
                'transaction_type': '입고',
                'transaction_date': receive_date if receive_date else None,
                'purpose': f'입고 성적서 업로드 (PO: {po_no})',
                'quantity': quantity,
                'lot_no': lot_no
            }).execute()
            print(f"raw_materials inserted (auto-id)")
        except Exception as e_raw:
            return jsonify({'status': 'error', 'message': f'입고 내역 저장 실패: {str(e_raw)}'}), 500

        return jsonify({'status': 'success', 'message': f'[{item_code}] {product_name} ({quantity} 입고) 내역이 정상 등록되었습니다.'})

    except Exception as e:
        import traceback
        err_msg = traceback.format_exc()
        print(f"Receiving Upload ERROR: {err_msg}")
        return jsonify({'status': 'error', 'message': str(e), 'trace': err_msg}), 500

@app.route('/api/usage/upload', methods=['POST'])
def upload_usage_api():
    if 'file' not in request.files:
        return jsonify({'status': 'error', 'message': '파일이 첨부되지 않았습니다.'}), 400

    file = request.files['file']
    if not file.filename or not file.filename.endswith('.xlsx'):
        return jsonify({'status': 'error', 'message': '.xlsx 파일만 업로드 가능합니다.'}), 400

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        if not rows:
            return jsonify({'status': 'error', 'message': '데이터가 없습니다. 2행부터 데이터를 입력해 주세요.'}), 400

        sb = get_supabase_client()
        success_count = 0
        errors = []
        
        insert_data = []

        for idx, row in enumerate(rows, start=2):
            try:
                item_code = str(row[0] or '').strip()
                product_name = str(row[1] or '').strip()
                lot_no = str(row[2] or '').strip()
                transaction_date = str(row[3] or '').strip()
                purpose = str(row[4] or '').strip()
                quantity = row[5]

                if not item_code or quantity is None:
                    errors.append(f'{idx}행: 원료코드 또는 사용량이 비어 있습니다.')
                    continue

                quantity = float(quantity)
                if quantity <= 0:
                    errors.append(f'{idx}행: 사용량은 0보다 커야 합니다. (값: {quantity})')
                    continue
                
                insert_data.append({
                    'product_name': product_name,
                    'item_code': item_code,
                    'lot_no': lot_no,
                    'transaction_type': '출고',
                    'transaction_date': transaction_date,
                    'purpose': purpose,
                    'quantity': quantity
                })
            except Exception as e:
                errors.append(f'{idx}행 파싱 오류: {str(e)}')
                
        if insert_data:
            try:
                # 500개씩 batch insert
                for i in range(0, len(insert_data), 500):
                    batch = insert_data[i:i+500]
                    sb.table('raw_materials').insert(batch).execute()
                    success_count += len(batch)
            except Exception as e:
                errors.append(f'DB 저장 중 오류: {str(e)}')

        return jsonify({
            'status': 'success',
            'success_count': success_count,
            'error_count': len(errors),
            'errors': errors
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'파일 처리 중 오류: {str(e)}'}), 500

@app.route('/api/usage/recent')
def get_recent_usage():
    try:
        sb = get_supabase_client()
        res = sb.table('raw_materials').select('id, item_code, product_name, lot_no, transaction_date, purpose, quantity').eq('transaction_type', '출고').order('id', desc=True).limit(50).execute()
        return jsonify({'status': 'success', 'data': res.data})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/finished-product/inventory')
def get_finished_product_inventory():
    try:
        sb = get_supabase_client()
        res = sb.table('finished_products').select('*').execute()
        
        # product_code, product_name, lot_no 복합키
        stats = defaultdict(lambda: {'mfg_date': None, 'exp_date': None, 'total_in': 0.0, 'total_out': 0.0})
        
        for r in res.data:
            key = (r['product_code'], r['product_name'], r['lot_no'])
            if r['transaction_type'] == '완제품입고':
                stats[key]['total_in'] += (r['quantity_kit'] or 0)
                if not stats[key]['mfg_date'] or r['transaction_date'] < stats[key]['mfg_date']:
                    stats[key]['mfg_date'] = r['transaction_date']
                if r.get('expire_date'):
                    stats[key]['exp_date'] = r['expire_date']
            elif r['transaction_type'] == '완제품 출고':
                stats[key]['total_out'] += (r['quantity_kit'] or 0)
                
        rows = []
        for (p_code, p_name, l_no), s in stats.items():
            current_stock = s['total_in'] - s['total_out']
            rows.append({
                'product_code': p_code,
                'product_name': p_name,
                'lot_no': l_no,
                'manufacture_date': s['mfg_date'],
                'expire_date': s['exp_date'],
                'total_in': s['total_in'],
                'total_out': s['total_out'],
                'current_stock': current_stock
            })
                
        rows.sort(key=lambda x: (x['manufacture_date'] or '', x['product_code'] or '', x['lot_no'] or ''))            
        return jsonify({'status': 'success', 'data': rows})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/finished-product/lot/<path:lot_no>')
def get_finished_product_lot_details(lot_no):
    try:
        sb = get_supabase_client()
        res = sb.table('finished_products').select('transaction_type, transaction_date, quantity_kit, destination, qc_info, remark').eq('lot_no', lot_no).order('transaction_date').order('id').execute()
        return jsonify({'status': 'success', 'data': res.data})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/finished-product/statistics')
def get_finished_product_statistics():
    year = request.args.get('year')
    product_code = request.args.get('product_code')
    
    try:
        sb = get_supabase_client()
        res = sb.table('finished_products').select('lot_no, product_code, transaction_type, transaction_date, quantity_kit').execute()
        
        mfg_dates = {}
        for r in res.data:
            if r['transaction_type'] == '완제품입고':
                key = (r['lot_no'], r['product_code'])
                if key not in mfg_dates or r['transaction_date'] < mfg_dates[key]:
                    mfg_dates[key] = r['transaction_date']
                    
        total_in = 0.0
        total_out = 0.0
        
        for r in res.data:
            key = (r['lot_no'], r['product_code'])
            mfg_date = mfg_dates.get(key)
            if not mfg_date:
                continue
                
            mfg_year = mfg_date[:4] if mfg_date else None
            
            if year and year != 'all' and mfg_year != year:
                continue
            if product_code and product_code != 'all' and r['product_code'] != product_code:
                continue
                
            if r['transaction_type'] == '완제품입고':
                total_in += (r['quantity_kit'] or 0)
            elif r['transaction_type'] == '완제품 출고':
                total_out += (r['quantity_kit'] or 0)
                
        stats = {
            'total_in': total_in,
            'total_out': total_out
        }
        return jsonify({'status': 'success', 'data': stats})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/raw-material/inventory')
def get_raw_material_inventory():
    """원료의 Lot별 상세 재고 및 정보 조회 (유효기간 포함)"""
    try:
        sb = get_supabase_client()
        
        # 1. 원료 마스터 정보 (material_info) 조회 및 중복 제거
        res_info = sb.table('material_info').select('*').execute()
        
        unique_info_map = {}
        for info in res_info.data:
            key = (info['item_code'], info['lot_no'])
            # 동일 (item_code, lot_no)가 있을 경우 입고일이 더 최근인 정보를 유지
            if key not in unique_info_map:
                unique_info_map[key] = info
            else:
                existing_date = unique_info_map[key].get('receive_date') or ''
                new_date = info.get('receive_date') or ''
                if new_date > existing_date:
                    unique_info_map[key] = info
        
        info_list = list(unique_info_map.values())
        
        # 2. 거래 내역 (raw_materials) 조회 및 Lot별 재고 계산
        res_raw = sb.table('raw_materials').select('item_code, lot_no, transaction_type, quantity').execute()
        
        stock_map = defaultdict(float) # (item_code, lot_no) -> current_stock
        for r in res_raw.data:
            key = (r['item_code'], r['lot_no'])
            qty = r['quantity'] or 0
            if r['transaction_type'] == '입고':
                stock_map[key] += qty
            else:
                stock_map[key] -= qty
        
        # 3. 데이터 결합
        result = []
        for info in info_list:
            key = (info['item_code'], info['lot_no'])
            current_stock = stock_map.get(key, 0.0)
            
            # 유효기간이 있는 항목만 처리하거나, 요청에 따라 모든 Lot 포함 가능
            # 여기서는 모든 원료 Lot 정보를 포함하도록 함
            result.append({
                'item_code': info['item_code'],
                'product_name': info['product_name'],
                'lot_no': info['lot_no'],
                'cat_no': info['cat_no'],
                'receive_date': info['receive_date'],
                'expire_date': info['expire_date'],
                'current_stock': round(current_stock, 4),
                'manufacturer': info['manufacturer'],
                'vendor': info['vendor']
            })
            
        # 정렬: 원료코드(item_code) 1순위, 유효기간(expire_date) 2순위
        # 유효기간이 없는 경우 정렬 시 아래로 가도록 '9999-12-31' 처리
        result.sort(key=lambda x: (x['item_code'] or '', x['expire_date'] or '9999-12-31'))
        return jsonify({'status': 'success', 'data': result})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/production/export-excel', methods=['POST'])
def export_production_excel():
    """생산 할당 내역을 엑셀로 추출합니다."""
    try:
        data = request.json
        items = data.get('items', [])
        usage_date = data.get('usage_date', '')
        usage_purpose = data.get('usage_purpose', '')
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "생산 할당 내역"
        
        # 헤더 설정
        headers = ["원료코드", "원료명", "할당 LOT", "할당 수량", "사용일자", "사용목적"]
        ws.append(headers)
        
        # 데이터 추가
        for item in items:
            mat_code = item.get('material_code', '')
            mat_name = item.get('material_name', '')
            
            allocated_lots = item.get('allocated_lots', [])
            if not allocated_lots:
                # 할당된 로트가 없는 경우 (재고 부족 등)
                ws.append([mat_code, mat_name, "할당 실패", 0, usage_date, usage_purpose])
            else:
                for lot in allocated_lots:
                    ws.append([
                        mat_code,
                        mat_name,
                        lot.get('lot_no', ''),
                        lot.get('allocated_qty', 0),
                        usage_date,
                        usage_purpose
                    ])
        
        # 스타일링 (헤더 배경색 등)
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        header_font = Font(bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # 컬럼 너비 자동 조정
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        file_date = usage_date.replace('-', '') if usage_date else datetime.datetime.now().strftime('%Y%m%d')
        filename = f"Production_Allocation_{file_date}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/bom-calculator')
def bom_calculator():
    return render_template('bom_calculator.html')

@app.route('/api/lots/bom')
def get_bom_lots():
    files = []
    bom_formulas_dir = find_data_file('bom_formulas')
    if os.path.exists(bom_formulas_dir):
        for filepath in glob.glob(os.path.join(bom_formulas_dir, "*.csv")):
            files.append({"filename": os.path.basename(filepath)})
            
    if not files:
        files = [{"filename": "BCE01_BOM_fomula_1.csv"}]
        
    return jsonify(files)


@app.route('/api/generate_bom', methods=['POST'])
def generate_bom():
    data = request.json
    target_qty = float(data.get('target_qty', 256))
    requested_file = data.get('formula_file', 'BCE01_BOM_fomula_1.csv')
    
    if target_qty <= 0:
        return jsonify({"error": "유효한 목표 생산량을 입력해주세요."}), 400
        
    try:
        safe_filename = os.path.basename(requested_file)
        target_path = os.path.join(find_data_file('bom_formulas'), safe_filename)
        
        if not os.path.exists(target_path):
            return jsonify({"error": f"파일 '{safe_filename}'을(를) 찾을 수 없습니다."}), 404
            
        df = None
        try:
            df = pd.read_csv(target_path, encoding='utf-8')
        except:
            df = pd.read_csv(target_path, encoding='cp949')
            
        if 'Level' in df.columns:
            df['Level'] = df['Level'].ffill()
            
        def evaluate_formula(formula_str, target_val):
            def excel_if(cond, t, f): return t if cond else f
            def roundup(n, digits=0):
                factor = 10**digits
                return math.ceil(n * factor) / factor

            allowed_names = {
                "target_qty": float(target_val), 
                "target_gty": float(target_val),
                "target": float(target_val),
                "round": round, "ROUND": round,
                "ROUNDUP": roundup, "ceil": math.ceil,
                "int": int, "float": float, "abs": abs, "ABS": abs,
                "IF": excel_if, "IFF": excel_if,
                "req": 1.0, "ratio": 1.0
            }
            try:
                f_str = str(formula_str).strip()
                if f_str.lower() == 'nan' or not f_str: return 0.0
                f_str = re.sub(r'\bif\s*\(', 'IFF(', f_str, flags=re.IGNORECASE)
                return float(eval(f_str, {"__builtins__": {}}, allowed_names))
            except Exception as e:
                return 0.0

        result = {
            "level0": {"제품명": f"전개 소스: {safe_filename}", "목표수량": target_qty},
            "level1": [], "level2": [], "level3": []
        }
        
        for _, r in df.iterrows():
            lvl_raw = r.get('Level')
            if pd.isna(lvl_raw): continue
            lvl = str(int(lvl_raw)) if isinstance(lvl_raw, (int, float)) else str(lvl_raw).strip()
            
            name1 = str(r.get('명칭 / 구성품', '')).strip()
            name2 = str(r.get('생산LOT', '')).strip()
            final_name = name2 if name2 and len(name2) > len(name1) and name2.lower() != 'nan' else name1
            
            formula = str(r.get('수식 (Formula)', '')).strip()
            unit = str(r.get('단위', '')).strip()
            parent = str(r.get('상위 LOT 연결', '')).strip()
            
            if not final_name or final_name.lower() in ['nan', '']: continue
            
            calculated_val = evaluate_formula(formula, target_qty)
            
            item_dict = {
                "상위Lot": parent if parent.lower() != 'nan' else '',
                "명칭 / 구성품": final_name,
                "생산Lot": name2 if name2.lower() != 'nan' else '',
                "계산된_소요량": round(calculated_val, 3),
                "단위": unit if unit.lower() != 'nan' else '',
                "레벨": lvl,
            }
            
            if lvl == '1': result["level1"].append(item_dict)
            elif lvl == '2': result["level2"].append(item_dict)
            elif lvl == '3': result["level3"].append(item_dict)
            
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/generate_bom_allocated', methods=['POST'])
def generate_bom_allocated():
    data = request.json
    target_qty = float(data.get('target_qty', 256))
    requested_file = data.get('formula_file', 'BCE01_BOM_fomula_1.csv')
    
    if target_qty <= 0:
        return jsonify({"error": "유효한 목표 생산량을 입력해주세요."}), 400
        
    try:
        safe_filename = os.path.basename(requested_file)
        target_path = os.path.join(find_data_file('bom_formulas'), safe_filename)
        
        if not os.path.exists(target_path):
            return jsonify({"error": f"파일 '{safe_filename}'을(를) 찾을 수 없습니다."}), 404
            
        df = None
        try:
            df = pd.read_csv(target_path, encoding='utf-8')
        except:
            df = pd.read_csv(target_path, encoding='cp949')
            
        if 'Level' in df.columns:
            df['Level'] = df['Level'].ffill()
            
        def evaluate_formula(formula_str, target_val):
            def excel_if(cond, t, f): return t if cond else f
            def roundup(n, digits=0):
                factor = 10**digits
                return math.ceil(n * factor) / factor

            allowed_names = {
                "target_qty": float(target_val), 
                "target_gty": float(target_val),
                "target": float(target_val),
                "round": round, "ROUND": round,
                "ROUNDUP": roundup, "ceil": math.ceil,
                "int": int, "float": float, "abs": abs, "ABS": abs,
                "IF": excel_if, "IFF": excel_if,
                "req": 1.0, "ratio": 1.0
            }
            try:
                f_str = str(formula_str).strip()
                if f_str.lower() == 'nan' or not f_str: return 0.0
                f_str = re.sub(r'\bif\s*\(', 'IFF(', f_str, flags=re.IGNORECASE)
                return float(eval(f_str, {"__builtins__": {}}, allowed_names))
            except Exception as e:
                return 0.0

        all_codes = set()
        for _, r in df.iterrows():
            name2 = str(r.get('생산LOT', '')).strip()
            if name2 and name2.lower() != 'nan':
                all_codes.add(name2)
        
        sb = get_supabase_client()
        lot_stocks = defaultdict(float)
        lot_in_qty = defaultdict(float)
        info_map = {}
        
        if all_codes:
            res_info = sb.table('material_info').select('item_code, lot_no, receive_date, expire_date').in_('item_code', list(all_codes)).execute()
            for r in res_info.data:
                info_map[(r['item_code'], r['lot_no'])] = {
                    'receive_date': r['receive_date'],
                    'expire_date': r['expire_date']
                }

            res_raw = sb.table('raw_materials').select('item_code, lot_no, transaction_type, quantity').in_('item_code', list(all_codes)).execute()
            for r in res_raw.data:
                key = (r['item_code'], r['lot_no'])
                qty = float(r['quantity'] or 0)
                if r['transaction_type'] == '입고':
                    lot_stocks[key] += qty
                    lot_in_qty[key] += qty
                else:
                    lot_stocks[key] -= qty

        result = {
            "level0": {"제품명": f"전개 소스: {safe_filename}", "목표수량": target_qty},
            "level1": [], "level2": [], "level3": []
        }
        
        for _, r in df.iterrows():
            lvl_raw = r.get('Level')
            if pd.isna(lvl_raw): continue
            lvl = str(int(lvl_raw)) if isinstance(lvl_raw, (int, float)) else str(lvl_raw).strip()
            
            name1 = str(r.get('명칭 / 구성품', '')).strip()
            name2 = str(r.get('생산LOT', '')).strip()
            final_name = name2 if name2 and len(name2) > len(name1) and name2.lower() != 'nan' else name1
            
            formula = str(r.get('수식 (Formula)', '')).strip()
            unit = str(r.get('단위', '')).strip()
            parent = str(r.get('상위 LOT 연결', '')).strip()
            
            if not final_name or final_name.lower() in ['nan', '']: continue
            
            calculated_val = evaluate_formula(formula, target_qty)
            mat_code = name2 if name2.lower() != 'nan' else ''
            required_qty = round(calculated_val, 3)
            
            allocated_lots = []
            shortage_qty = required_qty
            status = 'shortage'

            if mat_code and mat_code != '' and required_qty > 0:
                available_lots = []
                for (itm_cd, lot_no), stock in lot_stocks.items():
                    # 부동 소수점 오차를 고려하여 0.0001보다 큰 경우만 가용 재고로 인정
                    if itm_cd == mat_code and stock > 0.0001:
                        info = info_map.get((itm_cd, lot_no), {})
                        rec_date = info.get('receive_date') or '9999-12-31'
                        exp_date = info.get('expire_date') or '9999-12-31'
                        
                        total_in = lot_in_qty.get((itm_cd, lot_no), 0)
                        is_in_use = 0 if stock < total_in else 1
                        
                        available_lots.append({
                            'lot_no': lot_no,
                            'current_stock': stock,
                            'rec_date': rec_date,
                            'exp_date': exp_date,
                            'is_in_use': is_in_use
                        })
                
                available_lots.sort(key=lambda x: (x['is_in_use'], x['exp_date'], x['rec_date'], x['lot_no']))

                remaining_to_allocate = required_qty
                
                for lot in available_lots:
                    if remaining_to_allocate <= 0:
                        break
                        
                    alloc_qty = min(remaining_to_allocate, lot['current_stock'])
                    allocated_lots.append({
                        'lot_no': lot['lot_no'],
                        'receive_date': lot['rec_date'] if lot['rec_date'] != '9999-12-31' else '정보 없음',
                        'expire_date': lot['exp_date'] if lot['exp_date'] != '9999-12-31' else '정보 없음',
                        'available_stock': lot['current_stock'],
                        'allocated_qty': round(alloc_qty, 4)
                    })
                    
                    remaining_to_allocate -= alloc_qty
                    lot_stocks[(mat_code, lot['lot_no'])] = max(0, lot_stocks[(mat_code, lot['lot_no'])] - alloc_qty)

                shortage_qty = round(max(0, remaining_to_allocate), 4)
                status = 'success' if shortage_qty == 0 else 'shortage'
            elif required_qty == 0:
                shortage_qty = 0
                status = 'success'


            item_dict = {
                "상위Lot": parent if parent.lower() != 'nan' else '',
                "명칭 / 구성품": final_name,
                "Code_No": mat_code,
                "계산된_소요량": required_qty,
                "단위": unit if unit.lower() != 'nan' else '',
                "레벨": lvl,
                "allocated_lots": allocated_lots,
                "shortage_qty": shortage_qty,
                "status": status
            }
            
            if lvl == '1': result["level1"].append(item_dict)
            elif lvl == '2': result["level2"].append(item_dict)
            elif lvl == '3': result["level3"].append(item_dict)
            
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)

