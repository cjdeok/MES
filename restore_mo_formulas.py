import openpyxl

current_file = r'c:\Users\ENS-1000\Documents\Antigravity\MES\data\mo\MO_RESULT.xlsx'
original_file = r'c:\Users\ENS-1000\Documents\Antigravity\MES\.tmp\Manufacturing Order (MO)\MO_RESULT.xlsx'

def get_formulas(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    ws = wb.active
    formulas = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formulas[cell.coordinate] = cell.value
    return formulas

print("Extracting original formulas...")
orig_formulas = get_formulas(original_file)
print("Extracting current formulas...")
curr_formulas = get_formulas(current_file)

print("\nMissing or changed formulas (except E36):")
for coord, formula in orig_formulas.items():
    if coord == 'E36': continue
    if coord not in curr_formulas:
        print(f"MISSING: {coord} -> {formula}")
    elif curr_formulas[coord] != formula:
        print(f"CHANGED: {coord} -> Original: {formula}, Current: {curr_formulas[coord]}")

# If we found missing ones, let's restore them and update E36 correctly
if any(c != 'E36' for c in orig_formulas if c not in curr_formulas):
    print("\nRestoring formulas from original and applying new E36 formula...")
    wb = openpyxl.load_workbook(original_file, data_only=False)
    ws = wb.active
    new_e36_formula = "=ROUNDUP((IF($F$12=\"\",0,$F$12*6+40)+IF($F$13=\"\",0,$F$13*6+40)+IF($F$14=\"\",0,$F$14*6+40))+($F$15*20+30),-2)"
    ws['E36'] = new_e36_formula
    wb.save(current_file)
    print("Restore complete.")
else:
    print("\nNo other formulas were missing.")
