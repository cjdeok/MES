import openpyxl

orig_path = r'c:\Users\ENS-1000\Documents\Antigravity\MES\.tmp\Manufacturing Order (MO)\MO_RESULT.xlsx'
target_path = r'c:\Users\ENS-1000\Documents\Antigravity\MES\data\mo\MO_RESULT.xlsx'

# Load original
wb = openpyxl.load_workbook(orig_path, data_only=False)
ws = wb.active

# Apply new formula to E36
new_formula = "=ROUNDUP((IF($F$12=\"\",0,$F$12*6+40)+IF($F$13=\"\",0,$F$13*6+40)+IF($F$14=\"\",0,$F$14*6+40))+($F$15*20+30),-2)"
ws['E36'] = new_formula

# Count formulas
formulas = [c.coordinate for row in ws.iter_rows() for c in row if c.value and isinstance(c.value, str) and c.value.startswith('=')]
print(f"Total formulas in original (after E36 update): {len(formulas)}")

# Save to target
wb.save(target_path)
print("Successfully restored and updated template.")
